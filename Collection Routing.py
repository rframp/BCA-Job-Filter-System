import pandas as pd
import numpy as np
from ortools.constraint_solver import pywrapcp, routing_enums_pb2
from geopy.distance import geodesic
from openpyxl import load_workbook

# Load the data
file_path = r"C:\Users\RFrampton\OneDrive - Cowan Motor Group\Desktop\BCA Work.xlsx"
data = pd.read_excel(file_path, sheet_name="Scotland")
coll_coordinates = data[['JobNumber', 'CollLat', 'CollLon']].dropna()

# Define the starting location (Livingstone)
starting_location = (55.899819685016475, -3.5198384054833203)

# Combine starting location with job numbers and collection locations
coll_locations = [(None, starting_location[0], starting_location[1])] + list(coll_coordinates.to_records(index=False))

# Create the distance matrix for collection locations
def create_distance_matrix(locations):
    num_locations = len(locations)
    distance_matrix = np.zeros((num_locations, num_locations))
    for i in range(num_locations):
        for j in range(num_locations):
            if i != j:
                distance_matrix[i][j] = geodesic((locations[i][1], locations[i][2]), (locations[j][1], locations[j][2])).miles
    return distance_matrix

coll_distance_matrix = create_distance_matrix(coll_locations)

# Set up the VRP for collection data
def create_data_model(distance_matrix):
    data = {}
    data['distance_matrix'] = distance_matrix
    data['num_vehicles'] = int(np.ceil(len(distance_matrix) / 4))  # Total trips needed (each can carry 4)
    data['depot'] = 0  # Starting location index
    data['vehicle_capacities'] = [4] * data['num_vehicles']  # 4 locations per trip
    data['demands'] = [0] + [1] * (len(distance_matrix) - 1)  # Depot has demand 0, each location has demand 1
    return data

data = create_data_model(coll_distance_matrix)

# OR-Tools VRP Solver for collection data
def solve_vrp(data, locations):
    manager = pywrapcp.RoutingIndexManager(len(data['distance_matrix']), data['num_vehicles'], data['depot'])
    routing = pywrapcp.RoutingModel(manager)

    def distance_callback(from_index, to_index):
        from_node = manager.IndexToNode(from_index)
        to_node = manager.IndexToNode(to_index)
        return int(data['distance_matrix'][from_node][to_node] * 1000)  # Convert miles to integer scale

    transit_callback_index = routing.RegisterTransitCallback(distance_callback)
    routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)

    # Add capacity constraint
    def demand_callback(from_index):
        from_node = manager.IndexToNode(from_index)
        return data['demands'][from_node]

    demand_callback_index = routing.RegisterUnaryTransitCallback(demand_callback)
    routing.AddDimensionWithVehicleCapacity(
        demand_callback_index,
        0,  # no slack
        data['vehicle_capacities'],  # vehicle maximum capacities
        True,  # start cumul to zero
        'Capacity'
    )

    # Set search parameters
    search_parameters = pywrapcp.DefaultRoutingSearchParameters()
    search_parameters.first_solution_strategy = (
        routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC)
    search_parameters.local_search_metaheuristic = (
        routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH)
    search_parameters.time_limit.seconds = 30

    # Solve the problem
    solution = routing.SolveWithParameters(search_parameters)

    if solution:
        routes = []
        for vehicle_id in range(data['num_vehicles']):
            index = routing.Start(vehicle_id)
            route = []
            total_trip_distance = 0

            # Track previous location's coordinates for accurate segment distance
            prev_location = starting_location

            while not routing.IsEnd(index):
                node_index = manager.IndexToNode(index)
                if node_index != 0:  # Skip the depot itself
                    job_number, lat, lon = locations[node_index]
                    route.append((job_number, lat, lon))

                    # Calculate distance from previous point to the current one
                    segment_distance = geodesic(prev_location, (lat, lon)).miles
                    total_trip_distance += segment_distance
                    prev_location = (lat, lon)  # Update previous location

                index = solution.Value(routing.NextVar(index))

            # Add distance back to the starting location for a round trip
            return_to_start_distance = geodesic(prev_location, starting_location).miles
            total_trip_distance += return_to_start_distance

            routes.append((route, total_trip_distance))
        return routes
    else:
        print("No solution found!")
        return None

coll_routes = solve_vrp(data, coll_locations)

# Save Collection Data Routes to Excel
if coll_routes:
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        coll_result_data = []
        for route_id, (trip, distance) in enumerate(coll_routes, start=1):
            for loc in trip:
                job_number, lat, lon = loc
                coll_result_data.append([route_id, job_number, lat, lon, distance if loc == trip[0] else ""])
            coll_result_data.append([None, "", "", "", ""])  # Blank row

        coll_results_df = pd.DataFrame(coll_result_data, columns=["RouteID", "JobNumber", "CollLatitude", "CollLongitude", "Round Trip Distance (miles)"])
        coll_results_df.to_excel(writer, sheet_name="Optimal Scotland Routes", index=False, startrow=0, startcol=0)

wb = load_workbook(file_path)
ws = wb["Optimal Scotland Routes"]

# Adjust columns for Collection data (Script 1)
if "CollLatitude" in [cell.value for cell in ws[1]]:
    for col in ws.iter_cols(min_col=1, max_col=5, min_row=1):
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2  # Add padding

wb.save(file_path)
print("Completed writing grouped trips for Collection data.")
