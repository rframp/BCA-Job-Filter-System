import pandas as pd
from openpyxl import load_workbook
import numpy as np
import re
import os
from tkinter import Tk, messagebox
from tkinter.filedialog import askopenfilename

# Hide the root window
Tk().withdraw()

# Prompt the user to select an Excel file
file_path = askopenfilename(
    title="Select Excel File",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)

# Check if a file was selected
if not file_path:
    messagebox.showinfo("No File Selected", "No file selected. Exiting...")
else:
    try:
        # Define the Scottish postcode pattern
        scottish_pattern = re.compile(r'^(AB|DD|EH|FK|G|HS|IV|KA|KW|KY|ML|PA|PH|TD|ZE)\d', re.IGNORECASE)

        # Load data from Sheet2
        data = pd.read_excel(file_path, sheet_name="Sheet2")

        # Filter rows where 'CollPostCode' matches the Scottish postcode pattern and 'Distance' is greater than 130
        scotland_data = data[(data['CollPostCode'].str.match(scottish_pattern, na=False)) & (data['Distance'] > 130)].copy()

        # Ensure that the date columns are formatted to 'dd/mm/yyyy' if they exist
        date_columns = ['StartDate', 'EndDate', 'AgreedDate']
        for column in date_columns:
            if column in scotland_data.columns:
                # Apply formatting only if the value is non-null and non-empty
                scotland_data[column] = scotland_data[column].apply(
                    lambda x: pd.to_datetime(x, dayfirst=True, errors='coerce').strftime('%d/%m/%Y') if pd.notnull(x) and x != '' else ''
                )

        # Write the filtered data to a new sheet named "Scotland" after deleting the old sheet, if it exists
        wb = load_workbook(file_path)
        if "Scotland" in wb.sheetnames:
            del wb["Scotland"]
        wb.save(file_path)

        # Write to the new "Scotland" sheet
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a") as writer:
            scotland_data.to_excel(writer, sheet_name="Scotland", index=False)

        # Reload the workbook to access the new "Scotland" sheet
        wb = load_workbook(file_path)
        ws = wb["Scotland"]

        # Auto-adjust column widths to fit the content
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max_length + 2  # Add padding

        # Save the workbook after adjusting column widths
        wb.save(file_path)

        messagebox.showinfo("Completed", "Scotland sheet has been created. Check the sheet in Excel.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")
