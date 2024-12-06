import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension
from scipy.optimize import minimize
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
import numpy as np
import time

# Define the file path
file_path = r"C:\Users\RFrampton\OneDrive - Cowan Motor Group\Desktop\BCA Work.xlsx"

# Load the "Scotland" sheet data
scotland_data = pd.read_excel(file_path, sheet_name="Scotland")

# Drop any rows with missing or null postcodes in 'CollPostCode' to avoid extra rows
scotland_data = scotland_data.dropna(subset=['CollPostCode'])

# Initialize geolocator with delay to avoid rate limits
geolocator = Nominatim(user_agent="location_optimizer")

# Function to get latitude and longitude for a given postcode
def geocode_postcode(postcode):
    try:
        location = geolocator.geocode(postcode)
        if location:
            return location.latitude, location.longitude
    except GeocoderTimedOut:
        time.sleep(1)
        return geocode_postcode(postcode)
    return None, None

# Apply geocoding to each postcode and add results to the DataFrame if they are missing
if 'Latitude' not in scotland_data.columns or 'Longitude' not in scotland_data.columns:
    scotland_data[['Latitude', 'Longitude']] = scotland_data['CollPostCode'].apply(
        lambda postcode: pd.Series(geocode_postcode(postcode))
    )

# Drop rows where geocoding failed to find a location (i.e., Latitude or Longitude is NaN)
scotland_data = scotland_data.dropna(subset=['Latitude', 'Longitude'])

# Perform optimization to find the central location
coordinates = scotland_data[['Latitude', 'Longitude']].values

def total_distance(coord):
    lat, lon = coord
    distances = np.sqrt((coordinates[:, 0] - lat)**2 + (coordinates[:, 1] - lon)**2)
    return distances.sum()

# Initial guess (average latitude and longitude)
initial_guess = [coordinates[:, 0].mean(), coordinates[:, 1].mean()]

# Minimize the total distance
result = minimize(total_distance, initial_guess, method='L-BFGS-B')
optimal_lat, optimal_lon = result.x

# Function to check if a location is on land
def is_location_on_land(lat, lon):
    try:
        location = geolocator.reverse((lat, lon), exactly_one=True)
        if location and 'water' not in location.address.lower():
            return True
    except GeocoderTimedOut:
        time.sleep(1)
        return is_location_on_land(lat, lon)
    return False

# If initial optimized location is in water, perform a grid search
if not is_location_on_land(optimal_lat, optimal_lon):
    print("Optimal location is in water. Searching for nearby land using a grid pattern...")
    
    # Define the search grid parameters
    step_size = 0.01  # Increment in degrees (approximately ~1 km per 0.01 degree)
    search_radius = 0.1  # Search radius in degrees (~10 km)
    found_land = False

    # Grid search around the initial location
    for lat_offset in np.arange(-search_radius, search_radius + step_size, step_size):
        for lon_offset in np.arange(-search_radius, search_radius + step_size, step_size):
            test_lat = optimal_lat + lat_offset
            test_lon = optimal_lon + lon_offset
            if is_location_on_land(test_lat, test_lon):
                optimal_lat, optimal_lon = test_lat, test_lon
                found_land = True
                print("Found optimal land location.")
                break
        if found_land:
            break

    if not found_land:
        print("No land found in the search radius.")

print(f"Optimal Location: Latitude = {optimal_lat}, Longitude = {optimal_lon}")

# Ensure no duplicates in the latitude and longitude columns before saving
scotland_data = scotland_data.drop_duplicates(subset=['Latitude', 'Longitude'])

# Load the existing workbook and Scotland sheet
wb = load_workbook(file_path)
if "Scotland" not in wb.sheetnames:
    wb.create_sheet("Scotland")
ws = wb["Scotland"]

# Write the DataFrame to the "Scotland" sheet, clearing the sheet first
ws.delete_rows(2, ws.max_row)  # Clear all rows except headers
for r_idx, row in enumerate(dataframe_to_rows(scotland_data, index=False, header=True), start=1):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Write the optimal latitude and longitude to cells V2 and W2
ws["V2"].value = optimal_lat
ws["W2"].value = optimal_lon

# Auto-adjust column widths to fit the content
for col in ws.columns:
    try:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2  # Add padding
    except ValueError:
        continue

# Save the workbook
wb.save(file_path)
print("Data written successfully, and columns auto-adjusted.")
