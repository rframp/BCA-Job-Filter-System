import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
import time

# Define the file path
file_path = r"C:\Users\RFrampton\OneDrive - Cowan Motor Group\Desktop\BCA Work.xlsx"

# Load the "Scotland" sheet data
scotland_data = pd.read_excel(file_path, sheet_name="Scotland")

# Drop any rows with missing or null postcodes in 'CollPostCode' and 'DelPostCode' to avoid extra rows
scotland_data = scotland_data.dropna(subset=['CollPostCode', 'DelPostCode'])

# Initialize geolocator with delay to avoid rate limits
geolocator = Nominatim(user_agent="location_optimizer")

# Function to get latitude and longitude for a given postcode
def geocode_postcode(postcode):
    try:
        location = geolocator.geocode(postcode)
        if location:
            return location.latitude, location.longitude
    except GeocoderTimedOut:
        time.sleep(1)
        return geocode_postcode(postcode)
    return None, None

# Apply geocoding to 'CollPostCode' and 'DelPostCode' and add results to the DataFrame if they are missing
if 'CollLat' not in scotland_data.columns or 'CollLon' not in scotland_data.columns:
    scotland_data[['CollLat', 'CollLon']] = scotland_data['CollPostCode'].apply(
        lambda postcode: pd.Series(geocode_postcode(postcode))
    )

if 'DelLat' not in scotland_data.columns or 'DelLon' not in scotland_data.columns:
    scotland_data[['DelLat', 'DelLon']] = scotland_data['DelPostCode'].apply(
        lambda postcode: pd.Series(geocode_postcode(postcode))
    )

# Drop rows where geocoding failed to find a location (i.e., any Lat or Lon column is NaN)
scotland_data = scotland_data.dropna(subset=['CollLat', 'CollLon', 'DelLat', 'DelLon'])

# Ensure no duplicates in the latitude and longitude columns before saving
scotland_data = scotland_data.drop_duplicates(subset=['CollLat', 'CollLon', 'DelLat', 'DelLon'])

# Load the existing workbook and Scotland sheet
wb = load_workbook(file_path)
if "Scotland" not in wb.sheetnames:
    wb.create_sheet("Scotland")
ws = wb["Scotland"]

# Write the DataFrame to the "Scotland" sheet, clearing the sheet first
ws.delete_rows(2, ws.max_row)  # Clear all rows except headers
for r_idx, row in enumerate(dataframe_to_rows(scotland_data, index=False, header=True), start=1):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Auto-adjust column widths to fit the content
for col in ws.columns:
    try:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2  # Add padding
    except ValueError:
        continue

# Save the workbook
wb.save(file_path)
print("Geocoordinates written successfully.")
